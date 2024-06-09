# reading and reading/writing data from and to excel in python
import openpyxl
from openpyxl import load_workbook

wb1 = openpyxl.load_workbook("data.xlsx")
ws1 = wb1['Sheet1']
s = ws1['B2'].value
print(s)
# OR

s = ws1.cell(row=3, column=3).value
print(s)

# OR
data = ws1['A1':'C5']
for i in range(len(data)):
    for j in range(len(data[0])):
        print((data[i][j]).value)

for a, b, c in data:
    print(str(a.value) + " " + str(b.value) + " " + str(c.value))

#Writing in excel
## file must not be open in excel or anywhere
wb2 = openpyxl.load_workbook(r"C:\Users\dines\Learn And Interview\Interview_question_answer\Day_two\data.xlsx")
ws2 = wb2["Sheet2"]
ws2['B3'].value = 5
wb2.save(r"C:\Users\dines\Learn And Interview\Interview_question_answer\Day_two\data.xlsx")

col_roll = ws1['B1'].value
ws1['D1'] = "Multipler"
for i in range(2, 6):
    print('D'+str(i))
    ws1[('D' + str(i))] = 2*(int((ws1['B'+str(i)]).value))
wb1.save("data.xlsx")
print(col_roll)