import logging
import openpyxl
import softest

class Utility(softest.TestCase):

    def read_excel(self, filename, sheet):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook[sheet]
        dataList = []
        row_count = worksheet.max_row
        col_count = worksheet.max_column
        for i in range(2, row_count + 1):
            each_row_data = []
            for j in range(1, col_count + 1):
                each_row_data.append(worksheet.cell(row=i, column=j).value)
            dataList.append(each_row_data)
        return dataList

    def custom_logger(self):
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)
        console_handler = logging.StreamHandler()  # Handler for console
        console_handler.setLevel(logging.DEBUG)  # setLevel to handler
        file_handler = logging.FileHandler(
            r'C:\Users\dines\Learn And Interview\Interview_question_answer\Project\Logs\logs.txt',
            mode="a")  # Handler for file
        file_handler.setLevel(logging.DEBUG)  # setLevel to handler
        console_formatter = logging.Formatter(fmt='%(filename)s - %(asctime)s - %(levelname)s - %(message)s',
                                              datefmt='%d-%m-%Y %H:%M:%S')
        file_formatter = logging.Formatter(fmt='%(filename)s - %(asctime)s - %(levelname)s - %(message)s',
                                           datefmt='%d-%m-%Y %H:%M:%S')
        console_handler.setFormatter(console_formatter)  # add Formatter to handler
        file_handler.setFormatter(file_formatter)  # add Formatter to handler
        logger.addHandler(console_handler)  # added handler to logger
        logger.addHandler(file_handler)  # added handler to logger
        return logger

    def assertListItemText(self, list1, value):
        for stop in list1:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertIn, value, stop.text)
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()